import os
import sys
import time
import json
import random
import datetime
from playwright.sync_api import sync_playwright

_log_callback = None

def log_message(msg, status="info"):
    """로컬 대시보드 서버와 연동하여 실시간 진행 상태와 로그를 전송"""
    log_data = {
        "timestamp": datetime.datetime.now().strftime("%H:%M:%S"),
        "status": status,
        "message": msg
    }
    if _log_callback:
        try:
            _log_callback(log_data)
        except Exception:
            pass
    else:
        print(json.dumps(log_data), flush=True)

def split_weeks(start_date_str, end_date_str):
    """시작일과 종료일을 주간(7일) 단위로 분할하여 관리용 주차 메타데이터 목록 생성"""
    start = datetime.datetime.strptime(start_date_str, "%Y-%m-%d")
    end = datetime.datetime.strptime(end_date_str, "%Y-%m-%d")
    
    weeks = []
    current = start
    week_idx = 1
    
    while current <= end:
        week_end = current + datetime.timedelta(days=6)
        if week_end > end:
            week_end = end
            
        weeks.append({
            "index": week_idx,
            "label": f"{current.strftime('%Y')}년 {week_idx:02d}주차",
            "start": current.strftime("%Y-%m-%d"),
            "end": week_end.strftime("%Y-%m-%d")
        })
        current = week_end + datetime.timedelta(days=1)
        week_idx += 1
        
    return weeks

def format_date_range(start_str, end_str):
    """시작일과 종료일을 포털이 요구하는 en-dash 형식인 YYYY.MM.DD. – YYYY.MM.DD. 포맷으로 가공"""
    s_dt = datetime.datetime.strptime(start_str, "%Y-%m-%d")
    e_dt = datetime.datetime.strptime(end_str, "%Y-%m-%d")
    s_formatted = s_dt.strftime("%Y.%m.%d.")
    e_formatted = e_dt.strftime("%Y.%m.%d.")
    return f"{s_formatted} \u2013 {e_formatted}"

def convert_xlsx_to_csv(xlsx_path, csv_path):
    """Excel (XLSX) 파일을 CSV 파일로 신속하고 안전하게 변환"""
    try:
        import openpyxl
        import csv
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        sh = wb.active
        if not sh:
            return False
            
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            for row in sh.iter_rows(values_only=True):
                clean_row = ["" if x is None else str(x) for x in row]
                writer.writerow(clean_row)
        return True
    except Exception as e:
        log_message(f"CSV 변환 오류 발생: {str(e)}", "warning")
        return False

def split_bulk_file_into_weeks(bulk_xlsx_path, farm_name, download_dir, target_weeks):
    """다운로드받은 1개월/대용량 벌크 엑셀 파일을 로컬 메모리 상에서 주차별로 정밀 분할 가공"""
    try:
        import openpyxl
        log_message(f" 로컬 주차별 분할 연산 진행 중... (파일: {os.path.basename(bulk_xlsx_path)})", "info")
        
        # 벌크 파일 오픈
        wb_bulk = openpyxl.load_workbook(bulk_xlsx_path, data_only=True)
        sh_bulk = wb_bulk.active
        if not sh_bulk:
            log_message("[!] 벌크 엑셀 시트를 로드하지 못했습니다.", "warning")
            return
            
        # 데이터 행 파싱 (헤더: index, name, value, date, unit)
        rows = list(sh_bulk.iter_rows(values_only=True))
        if len(rows) <= 1:
            log_message(" -> 벌크 데이터 행이 비어 있어 분할을 스킵합니다.", "info")
            return
            
        header = rows[0] # ('index', 'name', 'value', 'date', 'unit')
        data_rows = rows[1:]
        
        # 각 주차별로 데이터를 필터링하여 저장
        for week in target_weeks:
            w_start = datetime.datetime.strptime(week['start'], "%Y-%m-%d").date()
            w_end = datetime.datetime.strptime(week['end'], "%Y-%m-%d").date()
            
            week_filtered_rows = []
            for r in data_rows:
                # r[3] 은 date 문자열 (예: '2025-05-01 00:00')
                date_str = r[3]
                if not date_str:
                    continue
                try:
                    # 날짜 부분만 추출하여 비교 (YYYY-MM-DD)
                    row_date = datetime.datetime.strptime(date_str.split()[0], "%Y-%m-%d").date()
                    if w_start <= row_date <= w_end:
                        week_filtered_rows.append(r)
                except Exception:
                    continue
            
            if not week_filtered_rows:
                continue
                
            # 주차별 폴더 생성 대신 농가명별 폴더에 바로 저장 (평탄화)
            farm_dir = os.path.join(download_dir, farm_name.replace("/", "_"))
            os.makedirs(farm_dir, exist_ok=True)
            
            # 파일명 형식: {주차라벨}_{시작일}_to_{종료일}.xlsx / .csv
            xlsx_filename = f"{week['label']}_{week['start']}_to_{week['end']}.xlsx"
            csv_filename = f"{week['label']}_{week['start']}_to_{week['end']}.csv"
            xlsx_path = os.path.join(farm_dir, xlsx_filename)
            csv_path = os.path.join(farm_dir, csv_filename)
            
            # 주차별 엑셀 파일 작성
            wb_new = openpyxl.Workbook()
            sh_new = wb_new.active
            sh_new.title = "data"
            sh_new.append(header)
            
            # index 컬럼을 1부터 다시 부여하여 정렬 저장
            for idx, r in enumerate(week_filtered_rows, 1):
                new_row = [idx] + list(r[1:])
                sh_new.append(new_row)
                
            wb_new.save(xlsx_path)
            wb_new.close()
            
            # CSV로 즉시 변환
            convert_xlsx_to_csv(xlsx_path, csv_path)
            log_message(f"  -> 저장 완료: {xlsx_filename} (.xlsx / .csv)", "success")
            
        wb_bulk.close()
    except Exception as e:
        log_message(f"[!] 로컬 주차 분할 가공 중 에러 발생: {str(e)}", "error")

def check_query_limits(page):
    """
    조회 기간 초과 또는 데이터 과다로 인해 조회 차단 경고가 떴는지 감시하는 지능형 감지기.
    경고가 뜨거나 EXCEL 버튼이 비활성화(disabled)된 경우 True를 반환.
    """
    try:
        # 1. 화면에 '최대 31일', '기간을 단축', '10일 이하' 등의 경고 텍스트가 노출되는지 체크
        alert_selectors = [
            "text*=최대 31일",
            "text*=최대 10일",
            "text*=기간을 단축",
            "text*=데이터가 너무 많습니다",
            "div.MuiAlert-message"
        ]
        for sel in alert_selectors:
            loc = page.locator(sel)
            if loc.count() > 0 and loc.first.is_visible():
                txt = loc.first.text_content()
                if any(k in txt for k in ["31일", "10일", "단축", "데이터가 너무"]):
                    log_message(f"[WAF Alert] 포털 제한 감지: '{txt.strip()}'", "warning")
                    return True
                    
        # 2. '테이블' 탭에 진입했을 때 EXCEL 버튼이 있고, 그것이 disabled 상태인지 추가 검증
        excel_btn = page.locator("button:has-text('EXCEL')").first
        if excel_btn.is_visible() and excel_btn.is_disabled():
            # 옆에 경고 텍스트가 같이 있는지 확인
            warning_text = page.locator("text*=최대 31일").first
            if warning_text.is_visible():
                log_message(f"[WAF Alert] EXCEL 다운로드 제한 감지 (31일 한도 초과 등)", "warning")
                return True
                
    except Exception:
        pass
    return False

def run_automation(config, callback=None):
    global _log_callback
    _log_callback = callback
    target_url = config.get("target_url", "https://jebitna.agri.jeju.kr/")
    username = config.get("username", "")
    password = config.get("password", "")
    start_date = config.get("start_date", "")
    end_date = config.get("end_date", "")
    download_dir = os.path.abspath(config.get("download_dir", "./downloads"))
    
    os.makedirs(download_dir, exist_ok=True)
    
    # 쪼개기 대상 주차 메타데이터
    weeks = split_weeks(start_date, end_date)
    total_weeks = len(weeks)
    log_message(f"수집 대상 구간 계산 완료: {start_date} ~ {end_date} (총 {total_weeks}개 주차 생성)", "info")
    
    with sync_playwright() as p:
        # 백그라운드 실행을 위해 headless=True 로 세팅!!
        log_message("백그라운드 백엔드 크롬 브라우저(Headless)를 기동합니다. (PC 사용 가능)", "info")
        browser = p.chromium.launch(
            headless=True,
            args=[
                "--disable-blink-features=AutomationControlled"
            ]
        )
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 800}
        )
        page = context.new_page()
        page.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        
        # 1. 로그인
        log_message("포털 접속 및 로그인을 진행합니다...", "info")
        page.goto(target_url, wait_until="networkidle")
        time.sleep(2.0)
        page.locator("input[type='text']").first.fill(username)
        page.locator("input[type='password']").first.fill(password)
        page.locator("input[type='password']").first.press("Enter")
        time.sleep(8.0)
        
        # 2. 동시계열데이터 페이지 이동
        log_message("동시계열데이터 메뉴로 이동 중...", "info")
        try:
            page.locator("button:has(svg[data-testid='MenuIcon'])").first.click()
            time.sleep(1.5)
            page.locator("text=모니터링").first.click()
            time.sleep(2.0)
            page.locator("text=동시계열데이터").first.click()
            time.sleep(6.0)
        except Exception as e:
            log_message(f"메뉴 이동 오류: {str(e)}. 직접 클릭해 주세요.", "warning")
            time.sleep(10.0)
            
        # 3. 허락된 전체 농가 목록 동적 스캔
        log_message("전체 농가 목록 동적 스캔 중...", "info")
        farm_names = []
        try:
            farm_input = page.locator("label:has-text('농장')").locator("xpath=../div/input").first
            farm_input.click()
            time.sleep(2.5)
            
            options = page.locator("[role='option']").all()
            for opt in options:
                name = opt.text_content().strip()
                if name:
                    farm_names.append(name)
            
            if options:
                options[0].click()
                time.sleep(1.0)
            log_message(f"농가 스캔 완료: {', '.join(farm_names)} (총 {len(farm_names)}개 농가)", "info")
        except Exception as e:
            log_message(f"농가 스캔 실패: {str(e)}. 기본 선택 농가로 수집합니다.", "warning")
            farm_names = ["기본농가"]
            
        total_farms = len(farm_names)
        
        # 4. 농가별 수집 시작
        for f_idx, farm_name in enumerate(farm_names):
            # 중지 요청 감시
            try:
                import run_dashboard
                if _log_callback and not run_dashboard.process_running:
                    log_message("사용자에 의해 수집 프로세스가 중지되었습니다.", "error")
                    browser.close()
                    return
            except Exception:
                pass
                
            log_message(f"\n==================================================", "info")
            log_message(f"[{f_idx + 1}/{total_farms}] 농장 수집 가동: {farm_name}", "progress")
            log_message(f"==================================================", "info")
            
            if total_farms > 1 and farm_name != "기본농가":
                try:
                    farm_input = page.locator("label:has-text('농장')").locator("xpath=../div/input").first
                    farm_input.click()
                    time.sleep(1.5)
                    page.locator(f"[role='option']:has-text('{farm_name}')").first.click()
                    time.sleep(2.5)
                    log_message(f"농장 변경 완료 -> '{farm_name}'", "info")
                except Exception as e:
                    log_message(f"농장 변경 실패 ({farm_name}): {str(e)}. 건너뜁니다.", "error")
                    continue
            
            # 농가별로 전체 기간(start_date ~ end_date)을 동적 슬라이싱 기법으로 처리
            global_start = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
            global_end = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
            
            current_start = global_start
            
            # 현재 농가에 대한 수집 완료 진행률 계산용 변수
            total_days = (global_end - global_start).days + 1
            
            while current_start <= global_end:
                # 중지 요청 감시
                try:
                    import run_dashboard
                    if _log_callback and not run_dashboard.process_running:
                        log_message("사용자에 의해 수집 프로세스가 중지되었습니다.", "error")
                        browser.close()
                        return
                except Exception:
                    pass
                    
                # 4-1. 기본 설정 크기인 '28일(4주 벌크)' 단위 조회를 먼저 시도
                chunk_size = 28
                success = False
                
                while chunk_size >= 3: # 최소 3일 단위까지 축소 가능
                    current_end = current_start + datetime.timedelta(days=chunk_size - 1)
                    if current_end > global_end:
                        current_end = global_end
                        
                    s_str = current_start.strftime("%Y-%m-%d")
                    e_str = current_end.strftime("%Y-%m-%d")
                    
                    days_done = (current_start - global_start).days
                    progress_percent = int(((f_idx * total_days + days_done) / (total_farms * total_days)) * 100)
                    
                    log_message(f"  [{progress_percent}%] {s_str} ~ {e_str} 조회 시도... (단위: {chunk_size}일)", "progress")
                    
                    # 날짜 주입
                    date_range_str = format_date_range(s_str, e_str)
                    date_input = page.locator("input[placeholder*='YYYY.MM.DD.']").first
                    date_input.click()
                    page.keyboard.press("Meta+A" if sys.platform == "darwin" else "Control+A")
                    page.keyboard.press("Backspace")
                    time.sleep(0.5)
                    date_input.fill(date_range_str)
                    page.keyboard.press("Enter")
                    time.sleep(1.0)
                    
                    # 센서 전체 선택
                    page.locator("button:has-text('전체 선택')").first.click()
                    time.sleep(1.0)
                    
                    # 데이터 조회 클릭 및 렌더링 대기
                    page.locator("button:has-text('데이터 조회')").first.click()
                    time.sleep(10.0) # 차트/테이블 로딩 대기
                    
                    # 테이블 탭 전환
                    page.locator("text=테이블").first.click()
                    time.sleep(3.0)
                    
                    # 지능형 한도 차단 체크
                    limit_hit = check_query_limits(page)
                    
                    if limit_hit:
                        # 한도 제한이 감지되면 조회 기간을 즉시 절반으로 줄여서 다시 시도
                        new_chunk = chunk_size // 2
                        log_message(f"  ⚠️ 농가별 한도 차단 감지! 수집 단위를 축소하여 재시도합니다. ({chunk_size}일 -> {new_chunk}일)", "warning")
                        chunk_size = new_chunk
                        continue # 이너 while 루프를 통해 단축된 기간으로 재시도
                    
                    # 다운로드 진행
                    excel_btn = page.locator("button:has-text('EXCEL')").first
                    if excel_btn.is_visible() and not excel_btn.is_disabled():
                        try:
                            # 임시 파일 경로
                            temp_dir = os.path.abspath("./temp_downloads")
                            os.makedirs(temp_dir, exist_ok=True)
                            temp_path = os.path.join(temp_dir, f"temp_bulk_{farm_name.replace('/', '_')}.xlsx")
                            
                            # 다운로드 수신
                            with page.expect_download(timeout=25000) as download_info:
                                excel_btn.click()
                            download = download_info.value
                            download.save_as(temp_path)
                            
                            log_message(f"  -> 벌크 파일 다운로드 성공: {s_str} ~ {e_str} ({chunk_size}일치)", "success")
                            
                            # 로컬 주차별 분할 처리 가동
                            split_bulk_file_into_weeks(temp_path, farm_name, download_dir, weeks)
                            
                            # 임시 파일 제거
                            if os.path.exists(temp_path):
                                os.remove(temp_path)
                                
                            success = True
                            break # 다운로드 성공 시 이너 while 루프 탈출
                        except Exception as dl_err:
                            log_message(f"  [!] 다운로드 스트림 오류: {str(dl_err)}. 단위를 축소해 재시도합니다.", "warning")
                            chunk_size = chunk_size // 2
                    else:
                        # 데이터가 없는 경우 (EXCEL 버튼 미노출 또는 비활성화이나 경고는 없는 경우)
                        log_message(f"  -> [PASS] 해당 기간에는 데이터가 존재하지 않습니다.", "info")
                        success = True
                        break
                        
                # 이너 while 루프 종료 후 처리
                if success:
                    # 성공한 기간의 다음 날짜로 포인터 이동
                    current_start = current_end + datetime.timedelta(days=1)
                else:
                    log_message(f"  [!] {current_start.strftime('%Y-%m-%d')} 기준 수집 불가. 하루를 건너뜁니다.", "error")
                    current_start = current_start + datetime.timedelta(days=1)
                    
                # 서버 과부하 방지 및 차단 방지 지연시간
                time.sleep(random.uniform(2.5, 4.0))
                
        log_message("모든 농가 및 3개년 스케줄 데이터 자동 수집 완료!", "complete")
        time.sleep(3.0)
        browser.close()

if __name__ == "__main__":
    if len(sys.argv) > 1:
        try:
            config = json.loads(sys.argv[1])
            run_automation(config)
        except Exception as e:
            print(json.dumps({"status": "error", "message": f"설정 파싱 에러: {str(e)}"}))
    else:
        # 디버그 실행용 모드 (2025년 5월 안정 데이터 수집 테스트)
        start_date_val = "2025-05-01"
        end_date_val = "2025-05-28"
        debug_config = {
            "target_url": "https://jebitna.agri.jeju.kr/",
            "username": "",
            "password": "",
            "start_date": start_date_val,
            "end_date": end_date_val,
            "download_dir": "./downloads"
        }
        run_automation(debug_config)
